# -*- coding: utf-8 -*-
"""Builds Friday_Test_Cases_and_Automation_Coverage.xlsx — a single-file
list of ALL Fridai test case scenarios (both already-automated and
not-yet-automated), each tagged with an "Automated?" column and a pointer
to the real automation script where one exists. Structured as a real
Excel Table so new rows can be appended below and keep filters/formatting.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
OUT_PATH = r"G:\My Drive\Rali\Fridai\outputs\Friday\Friday_Test_Cases_and_Automation_Coverage.xlsx"

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
AUTOMATED_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
NOT_AUTOMATED_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER_TOP = Alignment(wrap_text=True, vertical="top", horizontal="center")

COLUMNS = [
    ("Test ID", 12),
    ("Guide Section", 20),
    ("Title", 34),
    ("Type", 14),
    ("Priority", 10),
    ("Preconditions", 30),
    ("Test Data", 28),
    ("Steps", 42),
    ("Expected Result", 34),
    ("Postconditions", 24),
    ("Automated?", 12),
    ("Automation Script", 40),
    ("Author", 12),
    ("Notes", 30),
]

# Each row: (id, section, title, type, priority, preconditions, test_data,
#            steps, expected, postconditions, automated, script, author, notes)
ROWS = [
    ("FRIDAI-TC-001", "Before You Start", "Fridai test environment is reachable",
     "Smoke", "P1",
     "None",
     "BASE_URL=https://app.fridai.pro",
     "1. Navigate to BASE_URL.",
     "Page responds with HTTP 200 and a non-empty <title>.",
     "None",
     "Yes", "tests/smoke/test_homepage_loads.py::test_homepage_loads", "QA (Fridai project)", ""),

    ("FRIDAI-TC-002", "Step 1 — Log In", "Valid login lands on Dashboard",
     "Smoke", "P1",
     "A valid Fridai account exists",
     "FRIDAI_TEST_USER / FRIDAI_TEST_PASSWORD",
     "1. Go to /login.\n2. Enter email + password.\n3. Click Sign in.",
     "Redirected to Dashboard; \"+ Add Widget\" button visible.",
     "User is authenticated",
     "Yes", "tests/smoke/test_login.py::test_login_lands_on_dashboard", "QA (Fridai project)", ""),

    ("FRIDAI-TC-003", "Step 1 — Log In", "Login fails with an incorrect password",
     "Negative", "P2",
     "A valid Fridai account exists",
     "Valid email + deliberately wrong password",
     "1. Go to /login.\n2. Enter valid email, wrong password.\n3. Click Sign in.",
     "Login is rejected with a visible error; user stays on /login.",
     "User is NOT authenticated",
     "No", "", "QA (Fridai project)", "Gap — no negative login test yet"),

    ("FRIDAI-TC-004", "Step 1 — Log In", "\"Forgot Password\" link is absent from login page",
     "Regression", "P4",
     "None",
     "N/A",
     "1. Go to /login.\n2. Inspect the page for a password-reset link.",
     "No \"Forgot Password\" link is present (confirmed live 2026-07-16 — older docs describe one that doesn't exist).",
     "None",
     "No", "", "QA (Fridai project)", "Documents a known doc-vs-reality gap; not yet asserted in code"),

    ("FRIDAI-TC-005", "Step 2 — Navigation", "Orders & Fulfillment is reachable via sidebar nav",
     "Smoke", "P1",
     "Logged in",
     "N/A",
     "1. From Dashboard, expand \"Orders & Fulfillment\".\n2. Click \"Orders\".",
     "URL becomes /orders-list and the Orders & Fulfillment page loads.",
     "None",
     "Yes", "tests/orders/test_view_orders.py::test_orders_page_reachable_via_nav", "QA (Fridai project)", "Direct /orders navigation is a known broken link — always go via nav"),

    ("FRIDAI-TC-006", "Step 2 — Navigation", "Switching the global warehouse selector re-scopes the current page",
     "Functional", "P1",
     "Logged in; account has 2+ warehouses",
     "Two seeded warehouse names",
     "1. On Dashboard or Orders, note current data.\n2. Switch the top-bar warehouse selector to a different warehouse.\n3. Observe the page.",
     "Dashboard widgets / Orders list refresh to show only the newly selected warehouse's data.",
     "Selector remains on the newly chosen warehouse",
     "No", "", "QA (Fridai project)", "Gap — only exercised indirectly inside FRIDAI-TC-024/025 as setup, never asserted as its own scenario"),

    ("FRIDAI-TC-007", "Step 2 — Dashboard", "Add Widget catalog shows General/Operations/Network categories",
     "Functional", "P3",
     "Logged in",
     "N/A",
     "1. On Dashboard, click \"+ Add Widget\".\n2. Inspect the categorized list and select different widget types.",
     "Catalog shows General (Sales, Best Selling Products, Returns Report), Operations (8 widgets), and Network (2 ADMIN-badged widgets); selecting one updates the description panel.",
     "Modal can be cancelled without adding a widget",
     "No", "", "QA (Fridai project)", "Gap — manually explored only 2026-08-12, no DashboardPage coverage"),

    ("FRIDAI-TC-008", "Step 2 — Dashboard", "Network widgets are gated to Administrators",
     "Functional", "P2",
     "Two accounts: Administrator and a non-admin role",
     "Non-admin test user",
     "1. Log in as a non-admin user.\n2. Open \"+ Add Widget\".\n3. Check for the Network category.",
     "Network category (Transfers In Flight, Inventory Imbalance) is hidden or disabled for non-admin users.",
     "None",
     "No", "", "QA (Fridai project)", "Gap — the ADMIN badge's actual enforcement was never verified, only that the label exists"),

    ("FRIDAI-TC-009", "Step 3 — Warehouses", "Create a warehouse with \"Web fulfilment allowed\" enabled",
     "Functional", "P2",
     "Logged in as Administrator",
     "Distinct warehouse name, Priority=5, Type=Distribution Center",
     "1. Settings > Warehouses > \"+ Add Warehouse\".\n2. Fill Name, Priority, Warehouse Type.\n3. Check \"Web fulfilment allowed\".\n4. Click Create Warehouse.",
     "Warehouse appears in the Warehouses list and in the global warehouse selector.",
     "New warehouse exists and is selectable",
     "Yes", "tests/settings/test_create_warehouse.py::test_create_warehouse_with_web_fulfilment_allowed", "QA (Fridai project)", ""),

    ("FRIDAI-TC-010", "Step 3 — Warehouses", "Add Warehouse form defaults match spec",
     "Regression", "P3",
     "Logged in as Administrator",
     "N/A (form not submitted)",
     "1. Settings > Warehouses > \"+ Add Warehouse\".\n2. Inspect the 5 checkboxes without changing anything.",
     "Fulfillment enabled, Receiving enabled, Active are checked; Web fulfilment allowed, Default network fallback are unchecked.",
     "Modal can be cancelled",
     "Yes", "tests/settings/test_create_warehouse.py::test_new_warehouse_defaults_web_fulfilment_allowed_unchecked", "QA (Fridai project)", ""),

    ("FRIDAI-TC-011", "Step 3 — Warehouses", "Priority field determines pick order across web-enabled warehouses",
     "Functional", "P1",
     "2+ warehouses with \"Web fulfilment allowed\" checked and different Priority values",
     "WH-A Priority=1, WH-B Priority=5, both stocked for the same SKU",
     "1. Simulate/trigger a web-sourced order allocation across both warehouses.\n2. Observe which warehouse is allocated from first.",
     "The lower-Priority warehouse (WH-A) is allocated from first.",
     "None",
     "No", "", "QA (Fridai project)", "Gap — blocked until there's a way to actually trigger web-order allocation (pre-Hemi-integration); field exists but its effect is unverified"),

    ("FRIDAI-TC-012", "Step 3 — Warehouses", "Duplicate default warehouse names are avoided",
     "UX / Regression", "P3",
     "Creating 2+ warehouses without setting a custom Name",
     "N/A",
     "1. Create a warehouse without typing a distinct Name.\n2. Repeat for a second warehouse.\n3. Compare names in the selector.",
     "Each warehouse has a distinct, identifiable name (real bug found 2026-08-11: multiple warehouses all read \"Second warehouse\").",
     "None",
     "No", "", "QA (Fridai project)", "Tracks a real UX issue found on the seeded test account; not yet reported to the dev team or asserted in code"),

    ("FRIDAI-TC-013", "Step 4 — Locations", "Create a location via the 5-step wizard",
     "Functional", "P1",
     "Logged in",
     "Zone/Aisle/Bay/Bin names, e.g. \"Main Zone\"/\"A1\"/\"B1\"/\"BIN01\"",
     "1. Inventory Management > Locations > Create Location.\n2. Create new Zone, Aisle, Bay.\n3. Enter Bin, accept auto Location Code.\n4. Review and click Create location.",
     "New location is created with the concatenated Location Code and is selectable in Create Purchase Order's Receiving Location field.",
     "Location exists",
     "No", "", "QA (Fridai project)", "Gap — no LocationsPage Page Object exists yet"),

    ("FRIDAI-TC-014", "Step 4 — Locations", "Location Code can be overridden via \"Customize\"",
     "Functional", "P3",
     "Mid-way through the Create Location wizard, at the Bin step",
     "Custom code, e.g. \"CUSTOM-01\"",
     "1. At the Bin step, click \"Customize\".\n2. Enter a custom Location Code.\n3. Continue to Review and create.",
     "The location is created with the custom code, not the auto-generated concatenation.",
     "Location exists with custom code",
     "No", "", "QA (Fridai project)", "Gap"),

    ("FRIDAI-TC-015", "Step 5 — Suppliers", "Create a supplier with only the required field",
     "Functional", "P1",
     "Logged in",
     "Supplier Name only, e.g. \"Automation Test Supplier\"",
     "1. CRM & Suppliers > Suppliers > Add Supplier.\n2. Enter Name only.\n3. Click Create Supplier.",
     "Supplier is created and selectable in Create Purchase Order's Supplier field.",
     "Supplier exists",
     "No", "", "QA (Fridai project)", "Gap — no SuppliersPage; existing tests reuse the pre-seeded \"Rali test supplier\""),

    ("FRIDAI-TC-016", "Step 6 — Products", "Products page loads with core controls",
     "Smoke", "P1",
     "Logged in",
     "N/A",
     "1. Inventory Management > Products.",
     "\"Add Product\" button and the name/SKU/barcode search field are both visible.",
     "None",
     "Yes", "tests/products/test_view_products.py::test_products_page_loads", "QA (Fridai project)", ""),

    ("FRIDAI-TC-017", "Step 6 — Products", "A freshly created product has zero stock and is not orderable",
     "Regression / Negative", "P2",
     "Logged in",
     "New SKU/product name with a unique timestamp suffix",
     "1. Create a new Product (SKU + Name only).\n2. Start Create Order > Order Items.\n3. Search by the new SKU.",
     "No search result appears for the new SKU — it cannot be added to an order.",
     "No order/PO created",
     "Yes", "tests/orders/test_new_product_without_stock_not_orderable.py", "QA (Fridai project)", ""),

    ("FRIDAI-TC-018", "Step 7 — Get Stock (7a–7d)", "Full PO chain makes a product orderable",
     "Functional", "P1",
     "A supplier and a receiving location already exist",
     "New product, Supplier=\"Rali test supplier\", quantity=25",
     "1. Create Product.\n2. Create Purchase Order (Supplier, Location, status=Placed).\n3. Add PO line item, submit PO.\n4. Record ASN (Expected matches PO).\n5. Release for receiving.\n6. Receive: scan SKU, accept qty, scan location, Finish receiving.\n7. Search the SKU in Create Order.",
     "Product now appears in Order Items search with the received quantity available.",
     "Product has available stock",
     "Yes", "tests/inventory/test_receive_purchase_order.py::test_receiving_a_purchase_order_makes_a_product_orderable", "QA (Fridai project)", ""),

    ("FRIDAI-TC-019", "Step 7 — Get Stock", "Create Purchase Order's Warehouse field is disabled and pre-filled",
     "Regression", "P3",
     "Logged in; global warehouse selector set to a known warehouse",
     "N/A",
     "1. Purchase Orders > Create Purchase Order.\n2. Inspect the Warehouse field between Supplier and Receiving Location.",
     "Field is disabled and shows the same warehouse as the global selector.",
     "PO not submitted",
     "Yes (partial)", "pages/purchase_orders_page.py::current_warehouse()", "QA (Fridai project)", "Accessor exists and is exercised implicitly by TC-018; no dedicated assertion test yet"),

    ("FRIDAI-TC-020", "Step 7 — Cycle Count", "Cycle Count records stock without a Purchase Order",
     "Functional", "P2",
     "An existing product and location",
     "N/A",
     "1. Inventory Management > Cycle Count Tasks.\n2. Start a cycle count from Inventory or open an existing task.\n3. Record floor quantity.",
     "Product's available stock updates to reflect the counted quantity, without any PO involved.",
     "Stock updated",
     "No", "", "QA (Fridai project)", "Gap — documented via the app's own tour, never walked through even manually"),

    ("FRIDAI-TC-021", "Step 8 — Create Order", "Create and submit a new order (happy path)",
     "Functional", "P1",
     "At least one product with available stock (e.g. RaliPN1)",
     "Customer name/email, Channel=Direct to Consumer, SKU=RaliP1, qty=1",
     "1. Orders & Fulfillment > Create Order.\n2. Fill Customer Name/Email, choose Channel.\n3. Next > add line item > Create Order.",
     "New order appears in the Orders & Fulfillment list under the new customer's name.",
     "One new order exists",
     "Yes", "tests/orders/test_create_order.py::test_create_and_submit_order", "QA (Fridai project)", ""),

    ("FRIDAI-TC-022", "Step 8 — Create Order", "Order Items search returns nothing for an unknown SKU",
     "Negative", "P2",
     "Logged in",
     "Bogus SKU: \"ZZZ-DEFINITELY-NOT-A-REAL-PRODUCT-9999\"",
     "1. Start Create Order, fill Customer Info.\n2. In Order Items, search the bogus SKU.",
     "No search suggestion appears; nothing can be added.",
     "No order created",
     "Yes", "tests/orders/test_order_requires_product.py", "QA (Fridai project)", ""),

    ("FRIDAI-TC-023", "Step 9 — Orders list", "Packing progress is tracked separately from Status",
     "Regression", "P3",
     "Logged in",
     "N/A",
     "1. Go to Orders & Fulfillment.\n2. Inspect the table's column headers.",
     "Both a \"Packing\" column and a \"Status\" column are visible and independent.",
     "None",
     "Yes", "tests/orders/test_view_orders.py::test_orders_table_has_packing_column", "QA (Fridai project)", ""),

    ("FRIDAI-TC-024", "Step 9a — Pick", "Pick a picking task from allocated locations (full pick)",
     "Functional", "P1",
     "An order in \"Processing\" status with allocated stock",
     "Order from TC-021",
     "1. On the order's row, click \"Pick (N)\".\n2. Enter the full remaining quantity for each line (or click \"Mark All as Picked\").\n3. Click \"Complete Picking\".",
     "All allocated units are marked picked; order status becomes \"Picked\", next action becomes \"Pack (N)\".",
     "Order is ready to pack",
     "Yes", "tests/orders/test_partial_pick.py (via OrdersPage.click_next_action/set_pick_quantity/mark_all_as_picked/submit_complete_picking)", "QA (Fridai project)", "Implemented directly on OrdersPage (no separate PickingTasksPage) — the Pick modal opens inline from the Orders list, confirmed live 2026-08-20"),

    ("FRIDAI-TC-025", "Step 9b — Pack", "Pack a picked order without a shipping label",
     "Functional", "P2",
     "An order fully picked (see TC-024)",
     "N/A",
     "1. On the order's row, click \"Pack\".\n2. Pack the picked units into a box.\n3. Skip the shipping label step.",
     "Packing completes successfully — label is confirmed optional until a courier is configured.",
     "Order is ready to ship",
     "No", "", "QA (Fridai project)", "Gap — resolves a previously-flagged Bugs.docx item (\"Skip Label\" is OK) but was never turned into an automated check"),

    ("FRIDAI-TC-026", "Step 9c — Ship", "Ship a packed order and record the stock movement",
     "Functional", "P1",
     "An order fully packed (see TC-025)",
     "N/A",
     "1. On the order's row, click \"Ship\".\n2. Confirm shipment.",
     "Order status updates accordingly and a corresponding entry appears in Stock Movements.",
     "Order is shipped",
     "No", "", "QA (Fridai project)", "Gap — no ShippingTasksPage"),

    ("FRIDAI-TC-027", "Step 10a — Reallocate", "Reallocating an order moves it to a different warehouse",
     "Functional", "P1",
     "An existing order in a known warehouse; 2+ warehouses exist",
     "Order from TC-021; target = any other seeded warehouse",
     "1. Open the order > ORDER OPS > Reallocate.\n2. Choose a target warehouse.\n3. Click Reallocate.",
     "Order disappears from the origin warehouse's Orders list and appears under the target warehouse with the Warehouse column updated.",
     "Order now belongs to the target warehouse",
     "Yes", "tests/orders/test_reallocate_order.py::test_reallocate_order_moves_it_to_target_warehouse", "QA (Fridai project)", ""),

    ("FRIDAI-TC-028", "Step 10b — Split", "Splitting an order creates a sibling order in the same warehouse",
     "Functional", "P1",
     "An order with quantity > 1",
     "Order qty=10, split qty=3, target=\"Same as parent order\"",
     "1. Open the order > ORDER OPS > Split.\n2. Leave target warehouse as default.\n3. Enter quantity to move.\n4. Click Split order.",
     "A new sibling order is created (+1 total orders); parent's allocated qty drops to 7/7, sibling's is 3/3.",
     "Two independent orders now exist",
     "Yes", "tests/orders/test_split_order.py::test_split_order_creates_sibling_and_reduces_parent", "QA (Fridai project)", ""),

    ("FRIDAI-TC-029", "Step 10b — Split", "Splitting into a DIFFERENT target warehouse",
     "Functional", "P2",
     "An order with quantity > 1; 2+ warehouses exist",
     "Order qty=10, split qty=3, target=a different seeded warehouse",
     "1. Open the order > ORDER OPS > Split.\n2. Choose a different Target warehouse.\n3. Enter quantity, click Split order.",
     "Sibling order is created under the CHOSEN target warehouse, not the parent's warehouse.",
     "Sibling order belongs to the new target warehouse",
     "No", "", "QA (Fridai project)", "Gap — only the \"Same as parent order\" case is automated so far"),

    ("FRIDAI-TC-030", "Step 10b — Split", "Splitting more units than available is rejected",
     "Boundary / Negative", "P2",
     "An order with a known quantity, e.g. 10",
     "Split quantity = 11 (one more than the order has)",
     "1. Open the order > ORDER OPS > Split.\n2. Enter a quantity greater than what's on the order.\n3. Attempt to click Split order.",
     "Submission is blocked with a validation error; no sibling order is created.",
     "Order is unchanged",
     "No", "", "QA (Fridai project)", "Gap — boundary case never exercised"),

    ("FRIDAI-TC-031", "Step 10b — Split", "Splitting a multi-line order splits only the selected line",
     "Functional", "P2",
     "An order with 2+ distinct line items",
     "2-line order (2 different SKUs)",
     "1. Create an order with 2 line items.\n2. Open Split.\n3. Enter a quantity for only ONE line's row.\n4. Submit.",
     "Only the targeted line's units move to the sibling order; the other line stays entirely on the parent.",
     "Parent still has 1 full line + partial line; sibling has the split line only",
     "No", "", "QA (Fridai project)", "Gap — only single-line orders have been split so far; set_split_quantity(line_index=N) exists but untested"),

    ("FRIDAI-TC-032", "Step 11 — Stock Movements", "Every receive/pick/ship/count writes a movement",
     "Functional", "P2",
     "At least one completed receive and one shipped order",
     "N/A",
     "1. Inventory Management > Stock Movements.\n2. Filter/search for a known SKU or location.",
     "A movement entry exists for each of: the receive from TC-018, the pick/pack/ship from TC-024–026.",
     "None",
     "No", "", "QA (Fridai project)", "Gap — no Page Object; documented only via the app's own tour"),

    ("FRIDAI-TC-033", "Step 12 — Explore", "Stock Transfers moves stock between warehouses",
     "Functional", "P2",
     "2+ warehouses, one with available stock",
     "SKU with stock in WH-A, transfer to WH-B",
     "1. Inventory Management > Stock Transfers.\n2. Create a transfer: source WH-A, destination WH-B, SKU, quantity.\n3. Confirm.",
     "WH-A's available stock decreases and WH-B's increases by the transferred quantity.",
     "Stock relocated",
     "No", "", "QA (Fridai project)", "Gap — new nav item confirmed to exist 2026-08-11, never opened"),

    ("FRIDAI-TC-034", "Step 12 — Explore", "Standalone customer creation (not inline during order creation)",
     "Functional", "P3",
     "Logged in",
     "New customer name/email",
     "1. CRM & Suppliers > Customers > Add Customer.\n2. Fill required fields, save.\n3. Start Create Order and search for the new customer.",
     "New customer is created and selectable via \"Search Existing Customer\" in Create Order.",
     "Customer exists",
     "No", "", "QA (Fridai project)", "Gap — existing tests only use inline customer creation"),

    ("FRIDAI-TC-035", "Step 12 — Explore", "Returns processing for a shipped order",
     "Functional", "P2",
     "A shipped order (see TC-026)",
     "N/A",
     "1. Orders & Fulfillment > Returns.\n2. Start a return against the shipped order.\n3. Complete the return flow.",
     "Return is recorded and linked to the original order; returned stock is reflected appropriately.",
     "Return recorded",
     "No", "", "QA (Fridai project)", "Gap — Returns page never opened, not even manually"),

    ("FRIDAI-TC-036", "Step 8 — Create Order", "Web fulfilment-enabled warehouse is ready to receive a simulated web order",
     "Functional", "P1",
     "A warehouse with \"Web fulfilment allowed\" checked (see TC-009)",
     "N/A — no real Hemi integration exists yet",
     "1. Confirm the target warehouse has \"Web fulfilment allowed\" checked and an appropriate Priority.\n2. (Blocked) Trigger a web-sourced order once Hemi integration exists.",
     "Once integrated, a Hemi-sourced order is allocated to this warehouse ahead of lower-priority / non-web-enabled warehouses.",
     "N/A",
     "No", "", "QA (Fridai project)", "Blocked — cannot be exercised until Fridai<->Hemi integration exists; kept here so it isn't forgotten"),

    # ---- Aug 19, 2026 Order Queue visual rework (added 2026-08-20) ----

    ("FRIDAI-TC-037", "Step 9 — Order Queue", "Work Queues tabs filter the Orders table by lifecycle stage",
     "Functional", "P3",
     "Logged in; account has orders in multiple stages",
     "N/A",
     "1. Go to Orders.\n2. Click each Work Queues tab in turn: All orders, Ready to pick, Ready to pack, Ready to ship, Partially fulfilled, Backorders.",
     "Each tab filters the table to just that stage; the tab's count badge matches the number of rows shown.",
     "None",
     "No", "", "QA (Fridai project)", "Gap — OrdersPage.open_queue_tab()/queue_tab_count() exist but no test exercises clicking a tab and asserting the filtered result"),

    ("FRIDAI-TC-038", "Step 9 — Order Queue", "PROGRESS column reflects an order's pick/pack/ship stage",
     "Regression", "P2",
     "Logged in",
     "N/A",
     "1. Go to Orders.\n2. Inspect the PROGRESS column for orders at different stages.",
     "Shows a colored bar + text like \"5 ready to pack\"/\"3 ready to pick\", or plain \"Cancelled\" (no bar) for a cancelled order.",
     "None",
     "Yes", "tests/orders/test_partial_pick.py, test_split_order.py, test_combine_orders.py (via OrdersPage.order_row_progress_text)", "QA (Fridai project)", "Covered as an assertion inside other flows rather than a single dedicated test"),

    ("FRIDAI-TC-039", "Step 9 — Order Queue", "Row shows one dynamic next-action button + View + More actions",
     "Regression", "P2",
     "Logged in",
     "N/A",
     "1. Go to Orders.\n2. Inspect a row's ACTION column at different stages (e.g. Processing, Picked, Partially Picked).",
     "Exactly one bold next-action button appears (Pick (N)/Pack (N)/Ship (N)/\"X remaining (N)\"), plus \"View\", plus a \"More actions\" kebab when extra actions exist.",
     "None",
     "Yes", "tests/orders/test_partial_pick.py, test_cancel_the_rest (via OrdersPage.next_action_label/click_next_action)", "QA (Fridai project)", ""),

    ("FRIDAI-TC-040", "Step 9 — Order Queue", "\"More actions\" kebab reveals \"Cancel remaining\"",
     "Functional", "P3",
     "An order with units already picked/packed (partially fulfilled)",
     "Order from TC-043 (partial pick)",
     "1. On a partially-fulfilled order's row, click the \"More actions\" (⋯) button.\n2. Confirm \"Cancel remaining\" appears and is clickable.",
     "\"Cancel remaining\" is revealed as a real button; clicking it cancels the order's unfulfilled remainder.",
     "Order's remaining units are cancelled",
     "No", "", "QA (Fridai project)", "Gap — OrdersPage.open_more_actions()/click_more_action()/cancel_order() exist but no test calls them directly; only confirmed via manual/diagnostic investigation"),

    ("FRIDAI-TC-041", "Step 9 — Order Queue", "Bulk pick/pack/ship + Combine available with a specific warehouse active",
     "Functional", "P1",
     "A specific (non-\"All warehouses\") warehouse active in the global selector",
     "N/A",
     "1. Go to Orders with a specific warehouse active.\n2. Select the \"select all\" checkbox.\n3. Inspect the floating bulk action bar.",
     "\"Bulk pick (N)\", \"Bulk pack (N)\", \"Bulk ship (N)\", and \"Combine (N)\" are all visible.",
     "None",
     "Yes", "tests/orders/test_bulk_actions_warehouse_gating.py::test_bulk_pick_pack_ship_available_with_specific_warehouse", "QA (Fridai project)", ""),

    ("FRIDAI-TC-042", "Step 9 — Order Queue", "Bulk pick/pack/ship hidden under \"All warehouses\"; Allocate + Combine remain",
     "Functional", "P1",
     "\"All warehouses\" active in the global selector",
     "N/A",
     "1. Switch the global warehouse selector to \"All warehouses\".\n2. Select the \"select all\" checkbox.\n3. Inspect the floating bulk action bar.",
     "\"Bulk pick\"/\"Bulk pack\"/\"Bulk ship\" are absent; only \"Allocate (N)\" and \"Combine (N)\" remain.",
     "None",
     "Yes", "tests/orders/test_bulk_actions_warehouse_gating.py::test_bulk_pick_pack_ship_hidden_with_all_warehouses", "QA (Fridai project)", ""),

    ("FRIDAI-TC-043", "Step 9a — Pick", "Partial pick leaves the remainder on the same order",
     "Functional", "P1",
     "An order with quantity > 1",
     "Order qty=10, pick qty=4",
     "1. Click the order's \"Pick (N)\" button.\n2. Enter a quantity less than the full remaining amount.\n3. Click \"Complete Picking\".\n4. Choose \"Complete partial pick, leave rest as unpicked\".",
     "Order status becomes \"Partially Picked\"; progress reads \"4 ready to pack\"; next action becomes \"Pack remaining (4)\".",
     "Remainder (6 units) still pending, to be picked later",
     "Yes", "tests/orders/test_partial_pick.py::test_partial_pick_leaves_remainder_on_same_order", "QA (Fridai project)", ""),

    ("FRIDAI-TC-044", "Step 9a — Pick", "\"Complete Picking\" is disabled while every line reads 0",
     "Boundary / Negative", "P2",
     "A fresh, unpicked order",
     "Order qty=1",
     "1. Click the order's \"Pick (N)\" button.\n2. Leave the line's quantity at its default 0.\n3. Attempt to click \"Complete Picking\".",
     "\"Complete Picking\" button is disabled — cannot submit a pick of nothing.",
     "Order unchanged",
     "Yes", "tests/orders/test_partial_pick.py::test_complete_picking_disabled_at_zero_quantity", "QA (Fridai project)", ""),

    ("FRIDAI-TC-045", "Step 9a — Pick", "\"Cancel the rest\" marks the order fully \"Picked\", not \"Partially Picked\"",
     "Functional", "P1",
     "An order with quantity > 1",
     "Order qty=6, pick qty=2",
     "1. Click the order's \"Pick (N)\" button.\n2. Enter a partial quantity.\n3. Click \"Complete Picking\".\n4. Choose \"Complete partial pick, cancel the rest\".",
     "Order status becomes \"Picked\" (NOT \"Partially Picked\"); progress reads \"2 ready to pack\"; next action becomes plain \"Pack (2)\" (no \"remaining\").",
     "Cancelled units are gone from outstanding work; irreversible per the app's own warning",
     "Yes", "tests/orders/test_partial_pick.py::test_cancel_the_rest_marks_order_fully_picked", "QA (Fridai project)", "Confirmed live this is NOT equivalent to TC-043's outcome — the destructive choice fully closes out the order's Pick stage instead of leaving it partial"),

    ("FRIDAI-TC-046", "Step 9a — Pick", "Multi-line order: pick one line fully, another partially",
     "Functional", "P2",
     "An order with 2 distinct line items, both stocked",
     "Line A qty=5 (picked in full), Line B qty=8 (picked 3)",
     "1. Click the order's \"Pick (N)\" button.\n2. Locate each line by SKU (display order is not add-order).\n3. Enter 5 for Line A, 3 for Line B.\n4. Click \"Complete Picking\", choose \"leave rest as unpicked\".",
     "Order-level progress reflects the SUM across lines: \"8 ready to pack\"; next action \"Pack remaining (8)\" of 13 total items.",
     "Both lines' remaining units still pending",
     "Yes", "tests/orders/test_partial_pick.py::test_multiline_order_partial_pick_sums_across_lines", "QA (Fridai project)", "Uses OrdersPage.pick_line_index_for_sku() — the Pick modal's line display order does not match add-order"),

    ("FRIDAI-TC-047", "Step 10c — Combine", "Combine requires every selected order to share the same customer",
     "Functional / Negative", "P1",
     "2+ orders exist for the same customer; 2+ orders exist for different customers",
     "Same-customer pair; different-customer pair",
     "1. Select 2 orders for the SAME customer — confirm \"Combine\" is enabled.\n2. Clear selection, select 2 orders for DIFFERENT customers — confirm \"Combine\" is disabled.",
     "Combine is enabled only for the same-customer selection; disabled for the different-customer selection.",
     "No orders modified (precondition check only)",
     "Yes", "tests/orders/test_combine_orders.py::test_combine_requires_same_customer", "QA (Fridai project)", ""),

    ("FRIDAI-TC-048", "Step 10c — Combine", "Combine merges orders into one new order and cancels the sources",
     "Functional", "P1",
     "2 orders for the same customer, qty 2 and 3",
     "Source A qty=2, Source B qty=3",
     "1. Select both source orders.\n2. Click \"Combine\" in the bulk action bar.\n3. Confirm the \"Combine Orders\" modal lists both sources and warns sources will be cancelled.\n4. Click \"Combine orders\".",
     "Source orders become entirely unfindable (not just \"Cancelled\"); a new order appears tagged \"Combined\" with a \"from ORD-X, ORD-Y\" line and quantity 5 (2+3 summed).",
     "Sources gone; one new combined order in their place",
     "Yes", "tests/orders/test_combine_orders.py::test_combine_merges_orders_and_cancels_sources", "QA (Fridai project)", "Irreversible — same real-data caveat as other order-creation tests, but the cancellation here is intentional and permanent by design"),
]


def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # Header
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    automated_col = 11  # "Automated?" is column K
    for r_idx, row in enumerate(ROWS, start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.border = BORDER
            if c_idx in (5, 11):  # Priority, Automated? -> center
                cell.alignment = CENTER_TOP
            else:
                cell.alignment = WRAP_TOP
        automated_value = row[10]
        fill = AUTOMATED_FILL if automated_value.startswith("Yes") else NOT_AUTOMATED_FILL
        for c_idx in range(1, len(COLUMNS) + 1):
            ws.cell(row=r_idx, column=c_idx).fill = fill
        ws.row_dimensions[r_idx].height = 60

    last_row = len(ROWS) + 1
    last_col_letter = get_column_letter(len(COLUMNS))

    # Real Excel Table so new rows appended below keep formatting/filters.
    # Extend the table range a few rows below the current data so the user
    # can type new scenarios directly into a pre-formatted table row.
    EXTRA_BLANK_ROWS = 15
    table_last_row = last_row + EXTRA_BLANK_ROWS
    for r_idx in range(last_row + 1, table_last_row + 1):
        for c_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=r_idx, column=c_idx, value="")
            cell.font = Font(name=FONT_NAME, size=10)
            cell.border = BORDER
            cell.alignment = CENTER_TOP if c_idx in (5, 11) else WRAP_TOP

    table_ref = f"A1:{last_col_letter}{table_last_row}"
    table = Table(displayName="FridaiTestCases", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=False, showColumnStripes=False,
    )
    ws.add_table(table)

    # Dropdown validations for consistency on new rows too.
    dv_type = DataValidation(
        type="list",
        formula1='"Smoke,Regression,Functional,Negative,Boundary,UX,Negative / Boundary,UX / Regression,Functional / Negative,Boundary / Negative"',
        allow_blank=True,
    )
    dv_priority = DataValidation(type="list", formula1='"P1,P2,P3,P4"', allow_blank=True)
    dv_automated = DataValidation(type="list", formula1='"Yes,No,Yes (partial)"', allow_blank=True)
    for dv in (dv_type, dv_priority, dv_automated):
        ws.add_data_validation(dv)
    dv_type.add(f"D2:D{table_last_row}")
    dv_priority.add(f"E2:E{table_last_row}")
    dv_automated.add(f"K2:K{table_last_row}")

    ws.freeze_panes = "A2"

    # ---- Summary sheet ---------------------------------------------------
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 14

    ws2["A1"] = "Fridai Test Case Coverage — Summary"
    ws2["A1"].font = Font(name=FONT_NAME, bold=True, size=13)
    ws2.merge_cells("A1:B1")

    labels = [
        ("Total test cases", f"=COUNTA('Test Cases'!A2:A{table_last_row})"),
        ("Automated (Yes)", f"=COUNTIF('Test Cases'!K2:K{table_last_row},\"Yes\")"),
        ("Automated (Yes, partial)", f"=COUNTIF('Test Cases'!K2:K{table_last_row},\"Yes (partial)\")"),
        ("Not automated (No)", f"=COUNTIF('Test Cases'!K2:K{table_last_row},\"No\")"),
        ("P1 (Critical)", f"=COUNTIF('Test Cases'!E2:E{table_last_row},\"P1\")"),
        ("P2 (High)", f"=COUNTIF('Test Cases'!E2:E{table_last_row},\"P2\")"),
        ("P3 (Medium)", f"=COUNTIF('Test Cases'!E2:E{table_last_row},\"P3\")"),
        ("P4 (Low)", f"=COUNTIF('Test Cases'!E2:E{table_last_row},\"P4\")"),
    ]
    for i, (label, formula) in enumerate(labels, start=3):
        ws2.cell(row=i, column=1, value=label).font = Font(name=FONT_NAME, size=11)
        c = ws2.cell(row=i, column=2, value=formula)
        c.font = Font(name=FONT_NAME, size=11, bold=True)
        c.alignment = Alignment(horizontal="center")

    ws2["A11"] = "% Automated (incl. partial)"
    ws2["A11"].font = Font(name=FONT_NAME, size=11, italic=True)
    ws2["B11"] = "=(B4+B5)/B3"
    ws2["B11"].number_format = "0.0%"
    ws2["B11"].font = Font(name=FONT_NAME, size=11, bold=True)
    ws2["B11"].alignment = Alignment(horizontal="center")

    ws2["A13"] = "How to add a new test case scenario"
    ws2["A13"].font = Font(name=FONT_NAME, bold=True, size=11)
    ws2["A14"] = (
        "Go to the \"Test Cases\" tab and type into any of the pre-formatted "
        "blank rows already inside the table (or add a new row directly below "
        "the table — Excel/Sheets will auto-extend it). Use the next sequential "
        "FRIDAI-TC-### ID. The Type, Priority, and Automated? columns have "
        "dropdown validation built in."
    )
    ws2["A14"].alignment = Alignment(wrap_text=True, vertical="top")
    ws2.merge_cells("A14:B18")
    ws2.row_dimensions[14].height = 90

    wb.save(OUT_PATH)
    print("Wrote", OUT_PATH)


if __name__ == "__main__":
    build()
